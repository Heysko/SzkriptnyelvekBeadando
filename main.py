import requests
import tkinter as tk
import openpyxl
from openpyxl import Workbook, load_workbook
import os


# BookSearchApp osztály definiálása
class BookSearchApp:
    def __init__(self, root):
        # Inicializálja az alkalmazást
        self.root = root
        self.root.title("Könyv Kereső")

        self.search_label = tk.Label(root, text="Cím vagy Szerző:")
        self.search_label.pack()

        self.search_entry = tk.Entry(root)
        self.search_entry.pack()

        self.search_button = tk.Button(root, text="Keresés", command=self.search_books)
        self.search_button.pack()

        self.scroll_frame = tk.Frame(root)
        self.scroll_frame.pack()

        self.result_listbox = tk.Listbox(self.scroll_frame, width=50, height=15)
        self.result_listbox.pack(side=tk.LEFT)

        self.y_scrollbar = tk.Scrollbar(self.scroll_frame, orient=tk.VERTICAL, command=self.result_listbox.yview)
        self.y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.x_scrollbar = tk.Scrollbar(root, orient=tk.HORIZONTAL, command=self.result_listbox.xview)
        self.x_scrollbar.pack(fill=tk.X)

        self.result_listbox.config(yscrollcommand=self.y_scrollbar.set, xscrollcommand=self.x_scrollbar.set)

        self.detail_text = tk.Text(root, wrap=tk.WORD, width=50, height=10)
        self.detail_text.pack()

        # Listbox kijelölésének eseménykezelője
        self.result_listbox.bind('<<ListboxSelect>>', self.display_selected_book)

    # Találatok mentése fájlba
    def save_results_to_file(self, filename):
        selected_indices = self.result_listbox.curselection()
        if not os.path.exists(filename):
            # Ha a fájl nem létezik, létre hoz egy új munkafüzetet
            workbook = Workbook()
            worksheet = workbook.active

            # Adjuk hozzá a fejléceket, csak ha új fájlt hozunk létre
            worksheet.append(["Cím", "Szerző", "Kiadás éve", "Műfaj"])
            workbook.save(filename)
        else:
            # Be tölti  a már létező munkafüzetet
            workbook = load_workbook(filename)
            worksheet = workbook.active

            for index in selected_indices:
                selected_book = self.result_listbox.get(index)
                book_info = self.get_book_info(selected_book)

                if "Nincs találat" not in book_info:
                    worksheet.append([book_info["Cím"], book_info["Szerzők"], book_info["Kiadás éve"], book_info["Műfaj"]])

            # Menti a munkafüzetet
            workbook.save(filename)

    # Könyv keresése az Open Library API segítségével
    def search_books(self):
        query = self.search_entry.get()
        search_results = self.search_open_library(query)

        self.result_listbox.delete(0, tk.END)

        if search_results:
            for book in search_results:
                title = book.get('title', 'N/A')
                authors = ', '.join(book.get('author_name', ['N/A']))
                self.result_listbox.insert(tk.END, f"Cím: {title}, Szerző: {authors}")
        else:
            self.result_listbox.insert(tk.END, "Nincs találat")

    # Open Library API keresés
    def search_open_library(self, query):
        base_url = "http://openlibrary.org"
        endpoint = "/search.json"
        params = {
            'q': query
        }

        response = requests.get(base_url + endpoint, params=params)

        if response.status_code == 200:
            data = response.json()
            if 'docs' in data:
                return data['docs']
        return None

    # Kiválasztott könyv adatainak megjelenítése
    def display_selected_book(self, event):
        selected_indices = self.result_listbox.curselection()
        if selected_indices:
            index = int(selected_indices[0])
            selected_book = self.result_listbox.get(index)
            book_info = self.get_book_info(selected_book)
            self.detail_text.delete('1.0', tk.END)
            for key, value in book_info.items():
                self.detail_text.insert(tk.END, f"{key}: {value}\n")

    # Könyv adatainak lekérése az Open Library API segítségével
    def get_book_info(self, selected_book):
        if "Nincs találat" in selected_book:
            return {"Nincs találat": "Nincs találat a könyvre."}

        book_title = selected_book.split(",")[0].split(":")[1].strip()  # Könyvcím kinyerése
        base_url = "http://openlibrary.org"
        endpoint = f"/search.json?q={book_title}"

        response = requests.get(base_url + endpoint)

        if response.status_code == 200:
            data = response.json()
            if 'docs' in data and len(data['docs']) > 0:
                book = data['docs'][0]
                book_info = {
                    "Cím": book.get('title', 'N/A'),
                    "Szerzők": ', '.join(book.get('author_name', ['N/A'])),
                    "Kiadás éve": book.get('first_publish_year', 'N/A'),
                    "Műfaj": ', '.join(book.get('ol_genre', ['N/A'])),
                }
                return book_info
            else:
                return {"Nincs találat": "Nincs találat a könyvre."}
        return {"Hiba": "Nem sikerült lekérdezni az adatokat."}


# Excel fájl megnyitásának függvénye
def open_excel_file(file_path):
    try:
        os.startfile(file_path)
    except FileNotFoundError:
        print("Nem található a megadott fájl.")
    except Exception as e:
        print(f"Hiba történt: {e}")


# Excel fájlban lévő könyv adatok törlése
def clear_excel_data(file_path):
    excel_app = openpyxl.load_workbook(file_path)

    # Törli az összes sort a munkalapról, kivéve a fejléceket
    worksheet = excel_app.active
    for row in list(worksheet.iter_rows())[1:]:
        worksheet.delete_rows(row[1].row)
    excel_app.save(file_path)
    excel_app.close()


if __name__ == "__main__":
    root = tk.Tk()
    app = BookSearchApp(root)

    excel_file_path = "mentett_talalatok.xlsx"

    # Gombok hozzáadása az Excel fájl kezeléséhez
    save_button = tk.Button(root, text="Mentés", command=lambda: app.save_results_to_file("mentett_talalatok.xlsx"))
    save_button.pack()

    open_button = tk.Button(root, text="Excel Megnyitása", command=lambda: open_excel_file(excel_file_path))
    open_button.pack()

    clear_button = tk.Button(root, text="Könyv Adatok Törlése", command=lambda: clear_excel_data(excel_file_path))
    clear_button.pack()

    root.mainloop()
